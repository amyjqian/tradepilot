"""Fetch live SPDR sector ETF holdings from State Street.

State Street publishes daily holdings as XLSX files at:
  https://www.ssga.com/.../holdings-daily-us-en-{etf}.xlsx

Format: a few metadata rows (Fund Name / Ticker Symbol / "Holdings: As
of {date}"), a blank row, then the column header
  Name | Ticker | Identifier | SEDOL | Weight | Sector | Shares Held | Local Currency
followed by ~70 rows of holdings sorted by weight.

We parse the Ticker column, normalize symbols (BRK.B → BRK-B for
yfinance / Polygon compatibility), and persist to
`data_cache/sector_holdings.json` so subsequent scanner runs read from
disk instead of hitting SSGA on every request.

This module is opt-in: failures are non-fatal; `scanner.sector_rotation`
falls back to its hardcoded map when the JSON is absent or empty.
"""

from __future__ import annotations

import io
import json
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import httpx

log = logging.getLogger(__name__)


_SSGA_URL_TPL = (
    "https://www.ssga.com/us/en/individual/library-content/products/"
    "fund-data/etfs/us/holdings-daily-us-en-{etf}.xlsx"
)

# SSGA's CDN rejects the default httpx UA; a stock browser string passes.
_DEFAULT_UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0 Safari/537.36"
)


def _normalize_ticker(raw: Any) -> str | None:
    """Return a canonical ticker or None for non-equity rows.

    SSGA renders Berkshire as `BRK.B`; the rest of our codebase (and
    Polygon, yfinance) expects `BRK-B`. Cash/futures placeholders and
    blank rows are dropped.
    """
    if raw is None:
        return None
    s = str(raw).strip().upper()
    if not s or s in ("CASH_USD", "USD", "-", "N/A"):
        return None
    return s.replace(".", "-")


def fetch_ssga_holdings(etf: str, *, timeout: float = 15.0) -> list[str]:
    """Download + parse one SPDR ETF's holdings.

    Returns the ticker list in weight-descending order. Raises on HTTP
    failure or unparseable XLSX so the caller (`refresh_all_constituents`)
    can decide whether to skip just this ETF or abort the batch.
    """
    import openpyxl  # local import: parsing only happens during refresh

    url = _SSGA_URL_TPL.format(etf=etf.lower())
    resp = httpx.get(
        url,
        headers={"User-Agent": _DEFAULT_UA},
        follow_redirects=True,
        timeout=timeout,
    )
    resp.raise_for_status()

    wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError(f"SSGA workbook for {etf} has no active sheet")

    # Find the header row (cell value == "Ticker"). It's normally row 4
    # but SSGA has shuffled the leading metadata rows in past redesigns;
    # search the first 12 rows defensively.
    ticker_col_idx: int | None = None
    header_row_idx: int | None = None
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if r_idx > 12:
            break
        for c_idx, cell in enumerate(row):
            if cell and str(cell).strip().lower() == "ticker":
                ticker_col_idx = c_idx
                header_row_idx = r_idx
                break
        if ticker_col_idx is not None:
            break

    if ticker_col_idx is None or header_row_idx is None:
        raise ValueError(f"No 'Ticker' column found in SSGA XLSX for {etf}")

    tickers: list[str] = []
    seen: set[str] = set()
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if r_idx <= header_row_idx:
            continue
        if not row or len(row) <= ticker_col_idx:
            continue
        norm = _normalize_ticker(row[ticker_col_idx])
        if norm and norm not in seen:
            seen.add(norm)
            tickers.append(norm)
    return tickers


def refresh_all_constituents(etfs: list[str]) -> dict[str, list[str]]:
    """Fetch holdings for every ETF; per-ETF failures are logged and skipped."""
    out: dict[str, list[str]] = {}
    for etf in etfs:
        try:
            out[etf] = fetch_ssga_holdings(etf)
            log.info("SSGA: fetched %d holdings for %s", len(out[etf]), etf)
        except Exception as exc:
            log.warning("SSGA fetch failed for %s: %s", etf, exc)
    return out


def save_constituents(path: Path, holdings: dict[str, list[str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "fetched_at": datetime.now(timezone.utc).isoformat(),
        "source": "ssga",
        "holdings": holdings,
    }
    path.write_text(json.dumps(payload, indent=2))


def load_constituents(path: Path) -> tuple[dict[str, list[str]], str | None]:
    """Read the on-disk holdings cache.

    Returns `(holdings_map, fetched_at_iso)`. On any read or parse
    failure returns `({}, None)` so the caller falls back to its
    hardcoded source.
    """
    if not path.exists():
        return {}, None
    try:
        payload = json.loads(path.read_text())
    except Exception as exc:
        log.warning("Sector holdings cache %s unreadable: %s", path, exc)
        return {}, None
    holdings = payload.get("holdings") or {}
    if not isinstance(holdings, dict):
        return {}, None
    # Be defensive about value shape — only accept lists of strings.
    cleaned: dict[str, list[str]] = {}
    for etf, tickers in holdings.items():
        if isinstance(tickers, list):
            cleaned[str(etf)] = [str(t) for t in tickers if t]
    return cleaned, payload.get("fetched_at")
